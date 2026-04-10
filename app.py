from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import statistics
import textwrap

import openpyxl
import pandas as pd
import html
import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(
    page_title="Banking Analytics Prompt Library",
#     page_icon="📚",
#     layout="wide",
)
# ============================================================
# Data Models
# ============================================================

@dataclass
class PromptInput:
    """Represents an input field used by a prompt."""
    name: str
    description: str


@dataclass
class PromptRecord:
    """Represents one prompt entry in the library."""
    id: int
    name: str
    category: str
    tags: List[str]
    prompt_objective: str
    prompt_template: str
    required_inputs: List[PromptInput]
    optional_inputs: List[PromptInput]
    expected_result: str
    ratings: List[int] = field(default_factory=list)
    comments: List[str] = field(default_factory=list)


# ============================================================
# Excel Loader Utilities
# ============================================================

EXCEL_FILE_PATH = "prompt_guide.xlsx"
PROMPTS_SHEET_NAME = "Sheet2"   # change if your prompt master sheet has a different name
RATINGS_SHEET_NAME = "Ratings"
COMMENTS_SHEET_NAME = "Comments"

class PromptExcelLoader:
    """
    Converts the user's existing Excel table into PromptRecord objects.

    Expected columns:
    - Prompt ID
    - Prompt Name
    - Category
    - Tags/Labels
    - Prompt Objective
    - Prompt
    - Required Inputs
    - Sample Output

    Notes:
    - 'Category' is preserved because it matches the original sheet.
    - 'Required Inputs' can contain both required and optional fields.
    - 'Sample Output' may contain markdown and is rendered as markdown.
    """

    REQUIRED_COLUMNS = [
        "Prompt ID",
        "Prompt Name",
        "Category",
        "Tags/Labels",
        "Prompt Objective",
        "Prompt",
        "Required Inputs",
        "Sample Output",
    ]

    @classmethod
    def load_from_excel(cls, file_path: str) -> List[PromptRecord]:
        path = Path(file_path)
        if not path.exists():
            return []

        dataframe = pd.read_excel(path, sheet_name=PROMPTS_SHEET_NAME)
        dataframe.columns = [str(col).strip() for col in dataframe.columns]

        missing_columns = [col for col in cls.REQUIRED_COLUMNS if col not in dataframe.columns]
        if missing_columns:
            raise ValueError(f"Missing required Excel columns: {', '.join(missing_columns)}")

        records: List[PromptRecord] = []
        for _, row in dataframe.iterrows():
            record = cls._row_to_prompt_record(row)
            if record:
                records.append(record)
        return records

    @classmethod
    def _row_to_prompt_record(cls, row: pd.Series) -> Optional[PromptRecord]:
        prompt_id = cls._safe_int(row.get("Prompt ID"))
        prompt_name = cls._safe_text(row.get("Prompt Name"))

        if prompt_id is None or not prompt_name:
            return None

        category = cls._safe_text(row.get("Category")) or "Uncategorized"
        tags = cls._parse_tags(row.get("Tags/Labels"))
        prompt_objective = cls._safe_text(row.get("Prompt Objective"))
        prompt_template = cls._safe_text(row.get("Prompt"))
        expected_result = cls._safe_text(row.get("Sample Output"))

        required_inputs, optional_inputs = cls._parse_inputs(
            cls._safe_text(row.get("Required Inputs"))
        )

        return PromptRecord(
            id=prompt_id,
            name=prompt_name,
            category=category,
            tags=tags,
            prompt_objective=prompt_objective,
            prompt_template=prompt_template,
            required_inputs=required_inputs,
            optional_inputs=optional_inputs,
            expected_result=expected_result,
        )

    @staticmethod
    def _safe_text(value: Any) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip()

    @staticmethod
    def _safe_int(value: Any) -> Optional[int]:
        if pd.isna(value):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_tags(value: Any) -> List[str]:
        text = PromptExcelLoader._safe_text(value)
        if not text:
            return []

        normalized = text
        for sep in [",", ";", "|", "/"]:
            normalized = normalized.replace(sep, ",")

        return [item.strip() for item in normalized.split(",") if item.strip()]

    @staticmethod
    def _parse_inputs(value: str) -> Tuple[List[PromptInput], List[PromptInput]]:
        """
        Supports lightweight parsing from a single Excel cell.

        Accepted patterns include:
        - one item per line
        - comma-separated items
        - lines prefixed with 'Required:' or 'Optional:'
        - markdown bullets
        """
        if not value:
            return [], []

        required_inputs: List[PromptInput] = []
        optional_inputs: List[PromptInput] = []

        lines = [line.strip() for line in value.splitlines() if line.strip()]
        if not lines:
            lines = [item.strip() for item in value.split(",") if item.strip()]

        default_bucket = required_inputs

        for raw_line in lines:
            line = raw_line.lstrip("-*• ").strip()
            lower_line = line.lower()

            if lower_line.startswith("required:"):
                items = [item.strip() for item in line.split(":", 1)[1].split(",") if item.strip()]
                required_inputs.extend(
                    [PromptInput(name=item, description="Provided by user") for item in items]
                )
                default_bucket = required_inputs
                continue

            if lower_line.startswith("optional:"):
                items = [item.strip() for item in line.split(":", 1)[1].split(",") if item.strip()]
                optional_inputs.extend(
                    [PromptInput(name=item, description="Optional user input") for item in items]
                )
                default_bucket = optional_inputs
                continue

            default_bucket.append(PromptInput(name=line, description="Provided by user"))

        return required_inputs, optional_inputs

# ============================================================
# Feedback and Rating Layer
# ============================================================

def ensure_feedback_sheets_exist(file_path: str) -> None:
    """
    Creates Ratings and Comments sheets if they do not already exist.
    """
    path = Path(file_path)
    if not path.exists():
        return

    workbook = openpyxl.load_workbook(path)

    if RATINGS_SHEET_NAME not in workbook.sheetnames:
        rating_sheet = workbook.create_sheet(RATINGS_SHEET_NAME)
        rating_sheet.append(["Prompt ID", "Rating", "Created At"])

    if COMMENTS_SHEET_NAME not in workbook.sheetnames:
        comment_sheet = workbook.create_sheet(COMMENTS_SHEET_NAME)
        comment_sheet.append(["Prompt ID", "Comment", "Created At"])

    workbook.save(path)
    workbook.close()


def load_feedback_from_excel(file_path: str) -> Tuple[Dict[int, List[int]], Dict[int, List[str]]]:
    """
    Loads ratings and comments from Excel feedback sheets.
    """
    ratings_map: Dict[int, List[int]] = {}
    comments_map: Dict[int, List[str]] = {}

    path = Path(file_path)
    if not path.exists():
        return ratings_map, comments_map

    ensure_feedback_sheets_exist(file_path)

    try:
        ratings_df = pd.read_excel(path, sheet_name=RATINGS_SHEET_NAME)
        if not ratings_df.empty:
            for _, row in ratings_df.iterrows():
                prompt_id = row.get("Prompt ID")
                rating = row.get("Rating")
                if pd.notna(prompt_id) and pd.notna(rating):
                    pid = int(prompt_id)
                    ratings_map.setdefault(pid, []).append(int(rating))
    except Exception as exc:
        print(f"[WARN] Failed to load ratings: {exc}")

    try:
        comments_df = pd.read_excel(path, sheet_name=COMMENTS_SHEET_NAME)
        if not comments_df.empty:
            for _, row in comments_df.iterrows():
                prompt_id = row.get("Prompt ID")
                comment = row.get("Comment")
                if pd.notna(prompt_id) and pd.notna(comment):
                    pid = int(prompt_id)
                    comments_map.setdefault(pid, []).append(str(comment).strip())
    except Exception as exc:
        print(f"[WARN] Failed to load comments: {exc}")

    return ratings_map, comments_map


def append_rating_to_excel(file_path: str, prompt_id: int, rating: int) -> None:
    ensure_feedback_sheets_exist(file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[RATINGS_SHEET_NAME]
    sheet.append([prompt_id, rating, datetime.now().isoformat(timespec="seconds")])
    workbook.save(file_path)
    workbook.close()


def append_comment_to_excel(file_path: str, prompt_id: int, comment: str) -> None:
    ensure_feedback_sheets_exist(file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[COMMENTS_SHEET_NAME]
    sheet.append([prompt_id, comment, datetime.now().isoformat(timespec="seconds")])
    workbook.save(file_path)
    workbook.close()
    
# ============================================================
# Repository Layer
# ============================================================

class PromptRepository:
    """
    In-memory repository.

    This is intentionally separated from UI logic so the app can later be
    upgraded to use SQLite, Postgres, a REST API, or cloud storage.
    """

    def __init__(self, seed_data: Optional[List[PromptRecord]] = None):
        self._prompts: Dict[int, PromptRecord] = {}
        if seed_data:
            for prompt in seed_data:
                self._prompts[prompt.id] = prompt

    def list_all(self) -> List[PromptRecord]:
        return list(self._prompts.values())

    def get_by_id(self, prompt_id: int) -> Optional[PromptRecord]:
        return self._prompts.get(prompt_id)

    def search(self, search_text: str = "", category: str = "All") -> List[PromptRecord]:
        """
        Search by name, category, tags, objective, expected result, and prompt content.
        Also filter by category if provided.
        """
        search_text = (search_text or "").strip().lower()

        results = []
        for prompt in self._prompts.values():
            if category != "All" and prompt.category != category:
                continue

            haystack = " ".join([
                prompt.name,
                prompt.category,
                " ".join(prompt.tags),
                prompt.prompt_objective,
                prompt.expected_result,
                prompt.prompt_template,
            ]).lower()

            if not search_text or search_text in haystack:
                results.append(prompt)

        return sorted(results, key=lambda p: p.name.lower())

    def add_rating(self, prompt_id: int, rating: int) -> bool:
        prompt = self.get_by_id(prompt_id)
        if not prompt:
            return False

        prompt.ratings.append(rating)

        try:
            append_rating_to_excel(EXCEL_FILE_PATH, prompt_id, rating)
        except Exception as exc:
            print(f"[WARN] Failed to write rating to Excel: {exc}")

        return True

    def add_comment(self, prompt_id: int, comment: str) -> bool:
        prompt = self.get_by_id(prompt_id)
        if not prompt:
            return False

        prompt.comments.append(comment)

        try:
            append_comment_to_excel(EXCEL_FILE_PATH, prompt_id, comment)
        except Exception as exc:
            print(f"[WARN] Failed to write comment to Excel: {exc}")

        return True

    def categories(self) -> List[str]:
        categories = sorted({prompt.category for prompt in self._prompts.values()})
        return ["All"] + categories


# ============================================================
# Service Layer
# ============================================================

class PromptService:
    """
    Business logic layer.

    Keeps UI code thin and makes it easier to test or reuse elsewhere.
    """

    def __init__(self, repository: PromptRepository):
        self.repository = repository

    def get_prompt_choices(self, search_text: str, category: str) -> List[Tuple[str, int]]:
        prompts = self.repository.search(search_text, category)
        return [(f"{p.name} [{p.category}]", p.id) for p in prompts]

    def get_prompt_detail(self, prompt_id: int) -> Dict[str, str]:
        prompt = self.repository.get_by_id(prompt_id)
        if not prompt:
            return {
                "name": "",
                "category": "",
                "tags": "",
                "objective": "",
                "required_inputs": "",
                "optional_inputs": "",
                "expected_result": "",
                "rating_summary": "No ratings yet",
                "comments": "No comments yet",
                "visualizer": "",
                "copy_payload": "",
            }

        avg_rating = round(statistics.mean(prompt.ratings), 2) if prompt.ratings else None

        required_inputs = "\n".join(
            [f"• {inp.name}: {inp.description}" for inp in prompt.required_inputs]
        ) or "None"

        optional_inputs = "\n".join(
            [f"• {inp.name}: {inp.description}" for inp in prompt.optional_inputs]
        ) or "None"

        comments = "\n".join([f"• {c}" for c in prompt.comments]) or "No comments yet"

        rating_summary = (
            f"Average Rating: {avg_rating}/5 from {len(prompt.ratings)} review(s)"
            if avg_rating is not None
            else "No ratings yet"
        )

        visualizer = self.build_prompt_visualizer(prompt)
        copy_payload = self.build_copy_payload(prompt)

        return {
            "name": prompt.name,
            "category": prompt.category,
            "tags": ", ".join(prompt.tags),
            "objective": prompt.prompt_objective,
            "required_inputs": required_inputs,
            "optional_inputs": optional_inputs,
            "expected_result": prompt.expected_result,
            "rating_summary": rating_summary,
            "comments": comments,
            "visualizer": visualizer,
            "copy_payload": copy_payload,
        }

    def build_prompt_visualizer(self, prompt: PromptRecord) -> str:
        """
        Show only the prompt template in the view box.
        """
        return (prompt.prompt_template or "").strip()
    
    def build_copy_payload(self, prompt: PromptRecord) -> str:
        """Generates a copy-ready block that can be used directly by end users."""
        required_help = "\n".join(
            [f"• {inp.name}: {inp.description}" for inp in prompt.required_inputs]
        ) or "• None"

        optional_help = "\n".join(
            [f"• {inp.name}: {inp.description}" for inp in prompt.optional_inputs]
        ) or "• None"

        return textwrap.dedent(f"""
        Prompt ID: {prompt.id}
        Prompt Name: {prompt.name}
        Category: {prompt.category}
        Tags: {', '.join(prompt.tags)}
        Prompt Objective: {prompt.prompt_objective}

        Required Inputs:
        {required_help}

        Optional Inputs:
        {optional_help}

        Prompt Template:
        {prompt.prompt_template}

        Sample Output:
        {prompt.expected_result}
        """).strip()

    def submit_rating(self, prompt_id: int, rating: int) -> str:
        if not prompt_id:
            return "Please select a prompt before rating."
        if rating < 1 or rating > 5:
            return "Rating must be between 1 and 5."

        success = self.repository.add_rating(prompt_id, rating)
        return "Rating submitted successfully." if success else "Unable to submit rating."

    def submit_comment(self, prompt_id: int, comment: str) -> str:
        if not prompt_id:
            return "Please select a prompt before commenting."
        if not comment or not comment.strip():
            return "Comment cannot be empty."

        success = self.repository.add_comment(prompt_id, comment.strip())
        return "Comment added successfully." if success else "Unable to add comment."


# ============================================================
# Fallback Seed Data
# ============================================================





def render_copy_button(text: str, button_label: str, element_id: str) -> None:
    """
    Renders a clipboard copy button in Streamlit.
    After clicking, the label changes from 'Copy ...' to 'Copied'.
    """
    escaped_text = html.escape(text or "")
    escaped_label = html.escape(button_label)

    components.html(
        f"""
        <div style="margin: 0.25rem 0 1rem 0;">
            <button
                id="{element_id}"
                onclick="
                    navigator.clipboard.writeText(document.getElementById('{element_id}_payload').innerText);
                    const btn = document.getElementById('{element_id}');
                    const originalText = btn.innerText;
                    btn.innerText = 'Copied';
                    btn.disabled = true;
                    setTimeout(() => {{
                        btn.innerText = originalText;
                        btn.disabled = false;
                    }}, 1500);
                "
                style="
                    background: white;
                    border: 1px solid #d1d5db;
                    border-radius: 8px;
                    padding: 0.45rem 0.8rem;
                    font-size: 14px;
                    cursor: pointer;
                "
            >
                {escaped_label}
            </button>

            <div id="{element_id}_payload" style="display:none;">{escaped_text}</div>
        </div>
        """,
        height=55,
    )

def get_seed_prompts() -> List[PromptRecord]:
    """Sample prompts tailored for a banking analytics community."""
    return [
        PromptRecord(
            id=1,
            name="Credit Risk Summary Generator",
            category="Risk Analytics",
            tags=["credit-risk", "portfolio", "summary", "banking"],
            prompt_objective="Summarize credit portfolio performance and surface management actions.",
            prompt_template=(
                "You are a banking analytics expert. Analyze the following credit portfolio data: "
                "{portfolio_data}. Focus on delinquency trends, segment concentration, early warning signals, "
                "and recommended risk actions for management."
            ),
            required_inputs=[
                PromptInput("portfolio_data", "Portfolio snapshot, performance metrics, and delinquency measures."),
            ],
            optional_inputs=[
                PromptInput("time_period", "Reporting month, quarter, or year for comparison."),
                PromptInput("region", "Geography or branch segmentation if relevant."),
            ],
            expected_result="A concise risk summary with trends, red flags, and management recommendations.",
            ratings=[5, 4],
            comments=["Very useful for portfolio review meetings."],
        ),
        PromptRecord(
            id=2,
            name="Fraud Pattern Investigation Prompt",
            category="Fraud Analytics",
            tags=["fraud", "transactions", "anomaly", "investigation"],
            prompt_objective="Identify suspicious transaction patterns and prioritize fraud investigation actions.",
            prompt_template=(
                "Review the transaction anomaly data: {transaction_data}. Identify suspicious patterns, "
                "possible fraud typologies, customer segments impacted, and the likely next steps for investigation."
            ),
            required_inputs=[
                PromptInput("transaction_data", "Transaction records, anomaly scores, and alert details."),
            ],
            optional_inputs=[
                PromptInput("historical_baseline", "Normal transaction behavior benchmark."),
                PromptInput("channel", "ATM, card, mobile, branch, or online channel details."),
            ],
            expected_result="A structured fraud analysis highlighting patterns, risk severity, and investigation priorities.",
            ratings=[4, 4, 5],
            comments=["Good for preparing first-level fraud reviews."],
        ),
        PromptRecord(
            id=3,
            name="Customer Churn Insight Builder",
            category="Customer Analytics",
            tags=["churn", "retention", "segmentation", "customer"],
            prompt_objective="Explain attrition drivers and suggest segment-level retention actions.",
            prompt_template=(
                "Using this customer attrition dataset: {customer_data}, identify the top churn drivers, "
                "high-risk customer segments, and personalized retention strategies that a retail bank can deploy."
            ),
            required_inputs=[
                PromptInput("customer_data", "Customer profile, usage, product holding, and attrition indicators."),
            ],
            optional_inputs=[
                PromptInput("campaign_history", "Previous retention campaigns and outcomes."),
                PromptInput("lifecycle_stage", "Customer maturity stage such as new, growth, mature."),
            ],
            expected_result="A churn driver analysis with segment-level action recommendations.",
            ratings=[5],
            comments=["Can be extended with campaign personalization."],
        ),
        PromptRecord(
            id=4,
            name="Liquidity Dashboard Narrative",
            category="Treasury Analytics",
            tags=["liquidity", "treasury", "dashboard", "narrative"],
            prompt_objective="Convert treasury metrics into an executive liquidity narrative.",
            prompt_template=(
                "Turn the following treasury metrics into an executive narrative: {liquidity_metrics}. "
                "Comment on liquidity coverage, funding stability, short-term stress points, and actions needed."
            ),
            required_inputs=[
                PromptInput("liquidity_metrics", "Treasury and liquidity ratios, funding data, and scenario outputs."),
            ],
            optional_inputs=[
                PromptInput("stress_scenarios", "Scenario assumptions and stress-test outputs."),
            ],
            expected_result="An executive-ready treasury narrative summarizing liquidity position and concerns.",
            ratings=[],
            comments=[],
        ),
    ]


# ============================================================
# App Initialization Helpers
# ============================================================


def load_prompt_data() -> List[PromptRecord]:
    """
    Tries Excel first, then falls back to sample seed prompts.
    Also loads Ratings and Comments sheets from the same Excel file.
    """
    try:
        excel_records = PromptExcelLoader.load_from_excel(EXCEL_FILE_PATH)
        if excel_records:
            ensure_feedback_sheets_exist(EXCEL_FILE_PATH)
            ratings_map, comments_map = load_feedback_from_excel(EXCEL_FILE_PATH)

            for record in excel_records:
                record.ratings = ratings_map.get(record.id, [])
                record.comments = comments_map.get(record.id, [])

            return excel_records
    except Exception as exc:
        print(f"[WARN] Failed to load Excel data: {exc}")

    return get_seed_prompts()


@st.cache_resource
def get_app_service() -> PromptService:
    repository = PromptRepository(seed_data=load_prompt_data())
    return PromptService(repository)


# ============================================================
# Streamlit UI Helpers
# ============================================================


def init_session_state(service: PromptService) -> None:
    prompts = service.repository.list_all()
    first_id = prompts[0].id if prompts else None

    if "selected_prompt_id" not in st.session_state:
        st.session_state.selected_prompt_id = first_id


# ============================================================
# Streamlit UI
# ============================================================


st.markdown(
    """
    <style>
    .main .block-container {
        max-width: 1100px;
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .app-note {
        color: #6b7280;
        font-size: 0.95rem;
        margin-bottom: 1rem;
    }
    .field-label {
        font-size: 0.82rem;
        font-weight: 600;
        color: #6b7280;
        margin-bottom: 0.25rem;
    }
    .field-box {
        border: 1px solid #e5e7eb;
        border-radius: 8px;
        padding: 0.75rem 0.9rem;
        background: #ffffff;
        margin-bottom: 0.85rem;
        white-space: pre-wrap;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

service = get_app_service()
init_session_state(service)

st.title("Banking Analytics Prompt Library")

tab_library, tab_about = st.tabs(["Prompt Library", "Overview"])

with tab_about:
    st.subheader("Banking Analytics Prompt Library - Streamlit App")
    st.markdown(
        """
        ### Features
        - Interactive prompt visualizer
        - Prompt metadata:
          - prompt name
          - prompt category
          - relevancy tags
          - prompt objective
          - required inputs
          - optional inputs
          - expected result / sample output
        - User ratings
        - User comments
        - Copy-ready prompt output
        - Search bar
        - Category filter
        - Excel integration

        ### Design goals
        - Modular and scalable structure
        - Clear comments for maintainability
        - Easy to swap in persistent storage later (DB / API / file store)
        """
    )

with tab_library:
    st.markdown(
        '<div class="app-note">A clean, searchable library of reusable prompts for banking analytics.</div>',
        unsafe_allow_html=True,
    )

    search_col, filter_col = st.columns([3, 1])
    with search_col:
        search_text = st.text_input(
            "Search",
            placeholder="Search by prompt name, tag, category, or content",
        )
    with filter_col:
        category_filter = st.selectbox("Category", options=service.repository.categories(), index=0)

    filtered_prompts = service.repository.search(search_text=search_text, category=category_filter)
    prompt_choices = {f"{p.name} [{p.category}]": p.id for p in filtered_prompts}

    if not filtered_prompts:
        st.info("No prompts found for the current search/filter combination.")
        st.stop()

    current_ids = [p.id for p in filtered_prompts]
    if st.session_state.selected_prompt_id not in current_ids:
        st.session_state.selected_prompt_id = filtered_prompts[0].id

    selected_label = next(
        (label for label, pid in prompt_choices.items() if pid == st.session_state.selected_prompt_id),
        list(prompt_choices.keys())[0],
    )

    selected_prompt_label = st.selectbox(
        "Prompt",
        options=list(prompt_choices.keys()),
        index=list(prompt_choices.keys()).index(selected_label),
    )
    st.session_state.selected_prompt_id = prompt_choices[selected_prompt_label]

    selected_prompt_id = st.session_state.selected_prompt_id
    selected_prompt = service.repository.get_by_id(selected_prompt_id)
    detail = service.get_prompt_detail(selected_prompt_id)

    left_col, right_col = st.columns(2)

    with left_col:
        st.markdown('<div class="field-label">Name</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["name"]}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Category</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["category"]}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Tags</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["tags"] or "-"}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Prompt Objective</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["objective"] or "-"}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Required Inputs</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["required_inputs"]}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Optional Inputs</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="field-box">{detail["optional_inputs"]}</div>', unsafe_allow_html=True)

        st.markdown('<div class="field-label">Sample Output</div>', unsafe_allow_html=True)
        st.markdown('<div class="field-box">', unsafe_allow_html=True)
        st.markdown(detail["expected_result"] or "-")
        st.markdown('</div>', unsafe_allow_html=True)

    with right_col:
        st.markdown('<div class="field-label">Prompt View</div>', unsafe_allow_html=True)
        st.text_area(
            "Prompt View",
            value=detail["visualizer"],
            height=300,
            key="prompt_view_area",
            label_visibility="collapsed",
        )

        render_copy_button(
            text=detail["visualizer"],
            button_label="Copy View",
            element_id=f"copy_view_{selected_prompt_id}",
        )

        st.markdown('<div class="field-label">Prompt Template</div>', unsafe_allow_html=True)
        st.text_area(
            "Prompt Template",
            value=detail["copy_payload"],
            height=260,
            key="prompt_template_area",
            label_visibility="collapsed",
        )

        render_copy_button(
            text=detail["copy_payload"],
            button_label="Copy Prompt",
            element_id=f"copy_prompt_{selected_prompt_id}",
        )

    st.divider()
    feedback_col, comments_col = st.columns(2)

    with feedback_col:
        st.markdown("#### Ratings")
        st.caption(detail["rating_summary"])

        with st.form("rating_form", clear_on_submit=False):
            rating_value = st.slider("Rate Prompt", min_value=1, max_value=5, value=5, step=1)
            rating_submitted = st.form_submit_button("Submit Rating")
            if rating_submitted:
                message = service.submit_rating(selected_prompt_id, rating_value)
                get_app_service.clear()
                st.success(message)
                st.rerun()

    with comments_col:
        st.markdown("#### Comments")
        st.text_area(
            "Existing Comments",
            value=detail["comments"],
            height=160,
            disabled=True,
            label_visibility="collapsed",
        )

        with st.form("comment_form", clear_on_submit=True):
            comment_value = st.text_area(
                "Add Comment",
                placeholder="Share feedback or usage notes",
                height=100,
            )
            comment_submitted = st.form_submit_button("Add Comment")
            if comment_submitted:
                message = service.submit_comment(selected_prompt_id, comment_value)
                if "successfully" in message.lower():
                    get_app_service.clear()
                    st.success(message)
                    st.rerun()
                else:
                    st.warning(message)

    st.divider()
    with st.expander("Setup notes", expanded=False):
        st.markdown(
            """
            **Excel integration**
            - Place your Excel file in the same folder as this script.
            - File name expected: `prompt_library.xlsx`

            **Required Excel columns**
            - Prompt ID
            - Prompt Name
            - Category
            - Tags/Labels
            - Prompt Objective
            - Prompt
            - Required Inputs
            - Sample Output

            **Run locally**
            ```bash
            pip install streamlit pandas openpyxl
            streamlit run app.py
            ```
            """
        )